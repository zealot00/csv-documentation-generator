"""Excel document generator for CSV documentation"""

import os
from typing import Dict, Any, Optional, List
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import STANDARD_MODULES for RTM generation
import sys

sys.path.insert(0, str(Path(__file__).parent))
from requirements.parser import STANDARD_MODULES


class ExcelGenerator:
    """Generate Excel documents from templates"""

    # Default styles
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        doc_type: str,
        data: Optional[List[List[str]]] = None,
        headers: Optional[List[str]] = None,
        filename: Optional[str] = None,
        variables: Optional[Dict[str, str]] = None,
        db: Optional[Dict[str, Any]] = None,
    ) -> str:
        """Generate Excel document

        Args:
            doc_type: Type of document to generate
            data: Optional data rows for custom generation
            headers: Optional headers for custom generation
            filename: Optional custom filename
            variables: Optional template variables
            db: Optional requirements database for RTM generation
        """

        if variables is None:
            variables = {}

        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Set column widths for different document types
        self._set_column_widths(ws, doc_type)

        # Generate based on document type
        if doc_type == "rtm":
            self._generate_rtm(ws, variables, db)
        elif doc_type == "checklist":
            self._generate_checklist(ws, variables)
        elif doc_type == "test-case":
            self._generate_test_case(ws, variables)
        elif data and headers:
            self._generate_from_data(ws, headers, data)
        else:
            # Default template
            self._generate_default(ws, variables)

        # Save document
        if filename is None:
            project_name = variables.get("PROJECT_NAME", "")
            if project_name:
                filename = f"RTM_{project_name}.xlsx"
            else:
                filename = f"{doc_type.upper()}.xlsx"

        output_path = self.output_dir / filename
        wb.save(str(output_path))

        return str(output_path)

    def _set_column_widths(self, ws, doc_type: str):
        """Set column widths based on document type"""
        widths = {
            "rtm": [15, 35, 15, 10, 12, 12, 15, 12, 15, 12, 15, 12, 12, 20],
            "checklist": [30, 50, 20, 20],
            "test-case": [15, 30, 40, 20, 20, 20],
        }

        default_widths = [20, 40, 20, 20]

        col_widths = widths.get(doc_type, default_widths)

        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _generate_rtm(
        self, ws, variables: Dict[str, str], db: Optional[Dict[str, Any]] = None
    ):
        """Generate Traceability Matrix with 14 columns

        Columns:
        1. Requirement ID
        2. Requirement Description
        3. Module
        4. Priority
        5. FS Reference
        6. TS Reference
        7. IQ Test Case ID
        8. IQ Test Result
        9. OQ Test Case ID
        10. OQ Test Result
        11. PQ Test Case ID
        12. PQ Test Result
        13. Status
        14. Comments
        """
        # Headers
        headers = [
            "Requirement ID\n需求ID",
            "Requirement Description\n需求描述",
            "Module\n模块",
            "Priority\n优先级",
            "FS Reference\nFS参考",
            "TS Reference\nTS参考",
            "IQ Test Case ID\nIQ测试用例ID",
            "IQ Test Result\nIQ测试结果",
            "OQ Test Case ID\nOQ测试用例ID",
            "OQ Test Result\nOQ测试结果",
            "PQ Test Case ID\nPQ测试用例ID",
            "PQ Test Result\nPQ测试结果",
            "Status\n状态",
            "Comments\n备注",
        ]

        self._write_headers(ws, headers)

        # Get requirements and test results from db
        requirements = []
        test_results = []

        if db:
            requirements = db.get("requirements", [])
            test_results = db.get("test_results", [])

        # Create test results lookup
        test_results_by_req: Dict[str, Dict[str, str]] = {}
        for tr in test_results:
            req_id = tr.get("requirement_id")
            if req_id:
                if req_id not in test_results_by_req:
                    test_results_by_req[req_id] = {}
                test_type = tr.get("test_type", "OQ")
                test_results_by_req[req_id][test_type] = tr.get("status", "Pending")

        # Group requirements by module for organized output
        by_module: Dict[str, List[Dict]] = {}
        for req in requirements:
            module = req.get("module", "business_func")
            if module not in by_module:
                by_module[module] = []
            by_module[module].append(req)

        # Write requirements by module
        row = 2
        for module_id, module_reqs in by_module.items():
            # Get module name
            if module_id in STANDARD_MODULES:
                module_name = STANDARD_MODULES[module_id]["name_cn"]
            else:
                module_name = module_id

            for req in module_reqs:
                req_id = req.get("id", "")

                # Get test results for this requirement
                iq_result = test_results_by_req.get(req_id, {}).get("IQ", "Pending")
                oq_result = test_results_by_req.get(req_id, {}).get("OQ", "Pending")
                pq_result = test_results_by_req.get(req_id, {}).get("PQ", "Pending")

                row_data = [
                    req_id,
                    req.get("description", ""),
                    module_name,
                    req.get("priority", ""),
                    req.get("fs_ref", ""),
                    req.get("ts_ref", ""),
                    "",  # IQ Test Case ID - to be filled
                    iq_result,
                    "",  # OQ Test Case ID - to be filled
                    oq_result,
                    "",  # PQ Test Case ID - to be filled
                    pq_result,
                    req.get("status", "Draft"),
                    "",  # Comments
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = self.CELL_ALIGN
                    cell.border = self.BORDER

                row += 1

        # If no requirements, add sample rows
        if row == 2:
            for r in range(2, 12):
                for col in range(1, 15):
                    cell = ws.cell(row=r, column=col)
                    cell.alignment = self.CELL_ALIGN
                    cell.border = self.BORDER

    def _generate_checklist(self, ws, variables: Dict[str, str]):
        """Generate Validation Checklist"""
        headers = [
            "Check Item / 检查项",
            "Description / 描述",
            "Status / 状态",
            "Comments / 备注",
        ]

        self._write_headers(ws, headers)

        # Pre-defined checklist items for CSV
        checklist_items = [
            (
                "DOC-001",
                "User Requirements Specification (URS) prepared / 用户需求规格已准备",
                "",
            ),
            ("DOC-002", "Functional Specification (FS) prepared / 功能规格已准备", ""),
            ("DOC-003", "Risk Assessment completed / 风险评估已完成", ""),
            ("DOC-004", "IQ Protocol approved / IQ方案已批准", ""),
            ("DOC-005", "IQ executed and completed / IQ执行完成", ""),
            ("DOC-006", "OQ Protocol approved / OQ方案已批准", ""),
            ("DOC-007", "OQ executed and completed / OQ执行完成", ""),
            ("DOC-008", "PQ Protocol approved / PQ方案已批准", ""),
            ("DOC-009", "PQ executed and completed / PQ执行完成", ""),
            ("DOC-010", "Validation Summary Report approved / 验证总结报告已批准", ""),
            ("DOC-011", "Audit Trail reviewed / 审计追踪已审查", ""),
            ("DOC-012", "User training completed / 用户培训已完成", ""),
            (
                "DOC-013",
                "System maintenance procedure prepared / 系统维护规程已准备",
                "",
            ),
            ("DOC-014", "Backup and recovery tested / 备份恢复已测试", ""),
            ("DOC-015", "Change control procedure in place / 变更控制规程已建立", ""),
        ]

        for i, (item_id, description, comments) in enumerate(checklist_items, 2):
            ws.cell(row=i, column=1, value=item_id)
            ws.cell(row=i, column=2, value=description)
            ws.cell(row=i, column=3, value="")
            ws.cell(row=i, column=4, value=comments)

            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.alignment = self.CELL_ALIGN
                cell.border = self.BORDER

    def _generate_test_case(self, ws, variables: Dict[str, str]):
        """Generate Test Case Template"""
        headers = [
            "Test Case ID / 测试用例ID",
            "Test Scenario / 测试场景",
            "Test Steps / 测试步骤",
            "Expected Result / 预期结果",
            "Test Result / 测试结果",
            "Tester / 测试人",
        ]

        self._write_headers(ws, headers)

        # Sample rows
        for row in range(2, 12):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.alignment = self.CELL_ALIGN
                cell.border = self.BORDER

    def _generate_from_data(self, ws, headers: List[str], data: List[List[str]]):
        """Generate from provided data"""
        self._write_headers(ws, headers)

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = self.CELL_ALIGN
                cell.border = self.BORDER

    def _generate_default(self, ws, variables: Dict[str, str]):
        """Generate default template"""
        headers = ["Item", "Description", "Status", "Notes"]
        self._write_headers(ws, headers)

        # Sample data
        sample_data = [
            ["1", "Sample item 1", "Pending", ""],
            ["2", "Sample item 2", "Pending", ""],
            ["3", "Sample item 3", "Pending", ""],
        ]

        self._generate_from_data(ws, headers, sample_data)

    def _write_headers(self, ws, headers: List[str]):
        """Write header row"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.BORDER
